import json
import re
import base64
import mimetypes
import logging
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import load_workbook
import pdfplumber
from docx import Document
from anthropic import Anthropic
from config import ANTHROPIC_API_KEY, CLAUDE_MODEL, LOG_AI_PROMPTS, AI_PROMPT_LOG_MAX_CHARS
from PIL import Image
import io

try:
    import pytesseract
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False
    print("Warning: pytesseract not available. OCR will be disabled. Install tesseract-ocr for full functionality.")


class AIServiceError(Exception):
    """Raised when primary AI evaluation fails and caller should decide fallback behavior."""


logger = logging.getLogger(__name__)


class ComplianceEngine:
    def __init__(self, force_fallback: bool = False):
        self.force_fallback = force_fallback
        self.client = Anthropic(api_key=ANTHROPIC_API_KEY) if ANTHROPIC_API_KEY else None
        self.model = CLAUDE_MODEL
        self.ai_enabled = bool(ANTHROPIC_API_KEY) and not self.force_fallback
        self._vision_image_cache: Dict[str, Dict[str, str]] = {}
        if self.force_fallback:
            print("Info: Backup compliance evaluator enabled by request.")
        elif not self.ai_enabled:
            print("Warning: ANTHROPIC_API_KEY not set. Primary AI evaluation unavailable.")
        else:
            print(f"Info: Primary AI model: {self.model}")

    def _prompt_preview(self, prompt: str) -> str:
        prompt_text = (prompt or "").strip()
        if not prompt_text:
            return ""
        limit = max(200, int(AI_PROMPT_LOG_MAX_CHARS or 2000))
        if len(prompt_text) <= limit:
            return prompt_text
        return f"{prompt_text[:limit]}\n...[truncated {len(prompt_text) - limit} chars]"

    def _log_claude_prompt(self, mode: str, item: str, prompt: str) -> None:
        if not LOG_AI_PROMPTS:
            return
        item_snippet = " ".join(str(item or "").split())[:160]
        logger.info(
            "Claude prompt mode=%s model=%s item=%s chars=%s\n%s",
            mode,
            self.model,
            item_snippet,
            len(prompt or ""),
            self._prompt_preview(prompt),
        )

    def _build_json_prompt(self, payload: Dict[str, Any]) -> str:
        """Serialize model instructions as JSON text for deterministic prompt structure."""
        return json.dumps(payload, ensure_ascii=False)

    def parse_checklist(self, checklist_path: str) -> List[str]:
        """
        Parse checklist items from XLSX file.
        Supports sheets where column A is an index (e.g. "1.")
        and column B contains the actual checklist question.
        """
        def _strip_leading_serial(text: str) -> str:
            return re.sub(r"^\s*\d+\s*[\.\)\-:]\s*", "", (text or "")).strip()

        try:
            wb = load_workbook(checklist_path)
            ws = wb.active
            items = []

            # Identify best checklist text column from headers.
            label_col = None
            for c in range(1, ws.max_column + 1):
                header_val = ws.cell(1, c).value
                h = str(header_val).strip().lower() if header_val is not None else ""
                if any(k in h for k in ("label", "checklist", "question", "requirement")):
                    label_col = c
                    break

            if label_col is None and ws.max_column >= 2:
                label_col = 2

            for r in range(2, ws.max_row + 1):
                serial_raw = ws.cell(r, 1).value
                label_raw = ws.cell(r, label_col).value if label_col else None

                serial = str(serial_raw).strip() if serial_raw is not None else ""
                label = _strip_leading_serial(str(label_raw).strip()) if label_raw is not None else ""

                if not serial and not label:
                    continue

                if label:
                    if label.lower() in {"label", "response", "compliant / reject", "compliant/reject"}:
                        continue
                    # Do not prepend serial from column A.
                    # UI already has Sno/# column; checklist item should be clean text only.
                    items.append(label)
                    continue

                # Fallback to column A only when it's not just numeric index.
                if serial and not re.match(r"^\d+\.?$", serial):
                    items.append(_strip_leading_serial(serial))

            return items
        except Exception as e:
            print(f"Error parsing checklist: {e}")
            return []

    def extract_reference_text(self, reference_path: str) -> str:
        """Extract text from PDF or DOCX reference document."""
        if reference_path.endswith('.pdf'):
            return self._extract_pdf_text(reference_path)
        elif reference_path.endswith('.docx'):
            return self._extract_docx_text(reference_path)
        return ""

    def _extract_pdf_text(self, pdf_path: str) -> str:
        """Extract text from PDF using pdfplumber."""
        try:
            text = ""
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            print(f"Error extracting PDF text: {e}")
            return ""

    def _extract_docx_text(self, docx_path: str) -> str:
        """Extract text from DOCX using python-docx, preserving bold markers where possible."""
        try:
            doc = Document(docx_path)
            text = ""
            for para in doc.paragraphs:
                if para.runs:
                    parts = []
                    for run in para.runs:
                        run_text = run.text or ""
                        if not run_text:
                            continue
                        if run.bold:
                            parts.append(f"**{run_text}**")
                        else:
                            parts.append(run_text)
                    text += "".join(parts) + "\n"
                else:
                    text += para.text + "\n"
            return text
        except Exception as e:
            print(f"Error extracting DOCX text: {e}")
            return ""

    def extract_artwork_text(self, artwork_path: str) -> str:
        """Extract OCR text from artwork image."""
        try:
            if not TESSERACT_AVAILABLE:
                print(f"Warning: Tesseract not available. Install tesseract-ocr for OCR functionality.")
                return ""

            image = Image.open(artwork_path)
            # Use pytesseract for OCR
            text = pytesseract.image_to_string(image)
            return text
        except Exception as e:
            print(f"Error extracting artwork text: {e}")
            # Fallback: return empty string if pytesseract not available
            return ""

    def _evaluate_with_vision(self, item: str, reference_text: str, artwork_path: str) -> Dict[str, Any]:
        """
        Evaluate checklist item by examining the actual artwork image using Claude Vision API.
        """
        media_type, image_data = self._get_encoded_vision_image(artwork_path)

        prompt_payload = {
            "task": "packaging_compliance_check_from_image",
            "checklist_item": item,
            "reference_document_excerpt": (reference_text or "")[:3000],
            "instructions": [
                "Evaluate only this checklist row, not unrelated issues.",
                "If the row asks 'present', only judge presence.",
                "If the row asks 'according to reference', compare against reference text.",
                "For 'according to reference' rows, treat formatting as compliance-critical: spacing, punctuation, units, case, bold/emphasis, and symbols must match.",
                "If formatting cannot be clearly verified, default to Reject.",
                "Decision must be exactly one of: Compliant, Reject, N/A.",
            ],
            "output_schema": {
                "decision": "Compliant|Reject|N/A",
                "reference_info": "string",
                "artwork_info": "string",
                "comments": "string",
            },
            "output_requirements": {
                "format": "json_only",
                "no_extra_text": True,
            },
        }
        prompt = self._build_json_prompt(prompt_payload)

        self._log_claude_prompt("vision", item, prompt)

        last_parse_error = None
        for attempt in (1, 2):
            message = self.client.messages.create(
                model=self.model,
                max_tokens=1000,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": image_data,
                            },
                        },
                        {"type": "text", "text": prompt},
                    ],
                }],
            )

            raw_response = message.content[0].text if message.content else ""
            try:
                result = self._parse_model_response(
                    raw_response,
                    item,
                    allow_salvage=(attempt == 2),
                )
                result['checklist_item'] = item
                return result
            except json.JSONDecodeError as parse_err:
                last_parse_error = parse_err
                print(f"Warning: Vision JSON parse failed on attempt {attempt} for item '{item}': {parse_err}")
                if attempt == 1:
                    continue
                raise

        if last_parse_error:
            raise last_parse_error
        raise json.JSONDecodeError("Unable to parse vision response", "", 0)

    def _evaluate_product_name_with_vision(
        self,
        item: str,
        reference_text: str,
        artwork_path: str,
    ) -> Optional[Dict[str, Any]]:
        """
        Specialized product-name checker for stylized front-pack titles.
        Uses relaxed semantic matching (case/punctuation/line-break insensitive).
        """
        media_type, image_data = self._get_encoded_vision_image(artwork_path)
        ref_name = self._extract_reference_product_name(reference_text)
        item_l = re.sub(r"\s+", " ", (item or "").lower()).strip()
        check_mode = "presence_only" if ("present" in item_l and "according to reference" not in item_l) else "match_reference"

        prompt_payload = {
            "task": "product_name_compliance_from_image",
            "check_mode": check_mode,
            "checklist_item": item,
            "reference_product_name": ref_name,
            "reference_document_excerpt": (reference_text or "")[:1200],
            "instructions": [
                "Read the prominent product title on the front panel; ignore packed-for/contact address lines.",
                "Stylized decorative text counts as valid product-name evidence if clearly readable.",
                "For reference matching, compare semantic token equivalence; ignore case, punctuation, and line breaks.",
                "Do not fail solely due to OCR uncertainty when title is visually clear.",
                "Return only JSON.",
            ],
            "output_schema": {
                "decision": "Compliant|Reject|N/A",
                "reference_info": "string",
                "artwork_info": "string",
                "comments": "string",
            },
            "output_requirements": {"format": "json_only", "no_extra_text": True},
        }
        prompt = self._build_json_prompt(prompt_payload)
        self._log_claude_prompt("vision_product_name", item, prompt)

        try:
            message = self.client.messages.create(
                model=self.model,
                max_tokens=1000,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": image_data,
                            },
                        },
                        {"type": "text", "text": prompt},
                    ],
                }],
            )
            raw_response = message.content[0].text if message.content else ""
            parsed = self._parse_model_response(raw_response, item, allow_salvage=False)
            parsed["checklist_item"] = item
            return parsed
        except Exception as e:
            print(f"Warning: specialized product-name vision check failed for item '{item}': {e}")
            return None

    def _evaluate_allergen_format_with_vision(
        self,
        item: str,
        reference_text: str,
        artwork_path: str,
    ) -> Optional[Dict[str, Any]]:
        """
        Specialized strict allergen formatting parity check.
        Enforces bidirectional bold/unbold + spacing/punctuation comparison.
        """
        media_type, image_data = self._get_encoded_vision_image(artwork_path)
        allergen_terms = self._extract_allergen_terms(reference_text)
        prompt_payload = {
            "task": "allergen_format_parity_check_from_image",
            "checklist_item": item,
            "reference_allergen_terms": allergen_terms,
            "reference_document_excerpt": (reference_text or "")[:1800],
            "comparison_rules": [
                "Compare BOTH directions: reference->artwork and artwork->reference.",
                "For each allergen token (e.g., MILK, SOYA), verify bold/unbold formatting parity.",
                "Check spacing/punctuation parity for allergen statement formatting (e.g., colon/comma spacing).",
                "Any mismatch in bold/unbold/spacing/punctuation => Reject.",
                "If format cannot be verified confidently => Reject.",
                "Evaluate only the allergen row and ignore unrelated packaging issues.",
            ],
            "output_schema": {
                "decision": "Compliant|Reject|N/A",
                "reference_info": "string",
                "artwork_info": "string",
                "comments": "string",
            },
            "output_requirements": {"format": "json_only", "no_extra_text": True},
        }
        prompt = self._build_json_prompt(prompt_payload)
        self._log_claude_prompt("vision_allergen_format", item, prompt)

        try:
            message = self.client.messages.create(
                model=self.model,
                max_tokens=1000,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": image_data,
                            },
                        },
                        {"type": "text", "text": prompt},
                    ],
                }],
            )
            raw_response = message.content[0].text if message.content else ""
            parsed = self._parse_model_response(raw_response, item, allow_salvage=False)
            parsed["checklist_item"] = item
            return parsed
        except Exception as e:
            print(f"Warning: specialized allergen format vision check failed for item '{item}': {e}")
            return None

    def _get_encoded_vision_image(self, artwork_path: str) -> Tuple[str, str]:
        """
        Return (media_type, base64_data) for Vision input.
        Uses an in-memory cache so repeated row checks don't re-encode image bytes.
        """
        cached = self._vision_image_cache.get(artwork_path)
        if cached:
            return cached["media_type"], cached["data"]

        with open(artwork_path, "rb") as f:
            raw = f.read()

        try:
            img = Image.open(artwork_path)
            format_map = {
                "JPEG": "image/jpeg",
                "PNG": "image/png",
                "GIF": "image/gif",
                "WEBP": "image/webp",
            }
            media_type = format_map.get(img.format, "image/jpeg")
        except Exception:
            media_type = mimetypes.guess_type(artwork_path)[0] or "image/jpeg"

        encoded = base64.standard_b64encode(raw).decode("utf-8")
        self._vision_image_cache[artwork_path] = {"media_type": media_type, "data": encoded}
        return media_type, encoded

    def evaluate_checklist_item(self, item: str, reference_text: str, artwork_text: str, artwork_path: str = None) -> Dict[str, Any]:
        """
        Evaluate a single checklist item using Claude API.
        If artwork_path is provided, uses Vision API to examine the actual image.
        Otherwise falls back to OCR text analysis.
        Returns decision (Compliant/Reject/N/A) and comments.
        """
        if not self.ai_enabled or not self.client:
            return self._fallback_evaluation(item, reference_text, artwork_text)

        # Use vision-based evaluation if artwork_path is provided
        if artwork_path:
            try:
                result = self._evaluate_with_vision(item, reference_text, artwork_path)
                normalized_item = re.sub(r"\s+", " ", (item or "").lower()).strip()
                # Product-name rows get a dedicated relaxed semantic vision pass.
                if "product name" in normalized_item:
                    special = self._evaluate_product_name_with_vision(item, reference_text, artwork_path)
                    if special and special.get("decision") in {"Compliant", "Reject", "N/A"}:
                        result = special
                # Allergen strict row gets dedicated format-parity pass.
                if "allergen" in normalized_item and "according to reference" in normalized_item:
                    special_allergen = self._evaluate_allergen_format_with_vision(item, reference_text, artwork_path)
                    if special_allergen and special_allergen.get("decision") in {"Compliant", "Reject", "N/A"}:
                        result = special_allergen
                return self._apply_rule_overrides(item, reference_text, artwork_text, result)
            except Exception as e:
                print(f"Vision evaluation failed, falling back to text: {e}")
                # Fall through to text-based evaluation on failure

        try:
            prompt_payload = {
                "task": "packaging_compliance_check_from_text",
                "checklist_item": item,
                "reference_document_excerpt": (reference_text or "")[:2000],
                "artwork_detected_text_excerpt": (artwork_text or "")[:2000],
                "instructions": [
                    "Evaluate only this checklist row (do not reject for unrelated issues).",
                    "If the row asks whether something is 'present', focus on presence.",
                    "If the row asks whether content matches reference, compare with reference text.",
                    "For 'according to reference' rows, formatting is strict: spacing, punctuation, units, case, bold/emphasis, and symbols must match.",
                    "If formatting cannot be verified from provided text, choose Reject (strict mode).",
                ],
                "output_schema": {
                    "decision": "Compliant|Reject|N/A",
                    "reference_info": "string",
                    "artwork_info": "string",
                    "comments": "string",
                },
                "output_requirements": {
                    "format": "json_only",
                    "no_markdown": True,
                    "no_extra_text": True,
                },
            }
            prompt = self._build_json_prompt(prompt_payload)

            self._log_claude_prompt("text", item, prompt)

            last_parse_error = None
            for attempt in (1, 2):
                message = self.client.messages.create(
                    model=self.model,
                    max_tokens=1000,
                    messages=[
                        {"role": "user", "content": prompt}
                    ]
                )

                raw_response = message.content[0].text if message.content else ""
                try:
                    result = self._parse_model_response(
                        raw_response,
                        item,
                        allow_salvage=(attempt == 2),
                    )
                    result['checklist_item'] = item
                    return self._apply_rule_overrides(item, reference_text, artwork_text, result)
                except json.JSONDecodeError as parse_err:
                    last_parse_error = parse_err
                    print(f"Warning: Text JSON parse failed on attempt {attempt} for item '{item}': {parse_err}")
                    if attempt == 1:
                        continue
                    raise

            if last_parse_error:
                raise last_parse_error
            raise json.JSONDecodeError("Unable to parse text response", "", 0)

        except json.JSONDecodeError as e:
            print(f"Error parsing Claude response: {e}")
            fallback = {
                'checklist_item': item,
                'decision': 'N/A',
                'reference_info': 'Unable to evaluate',
                'artwork_info': 'Evaluation error',
                'comments': f'JSON parse error: {str(e)}'
            }
            return self._apply_rule_overrides(item, reference_text, artwork_text, fallback)
        except Exception as e:
            print(f"Error evaluating item: {e}")
            err_l = str(e).lower()
            if (
                "authentication_error" in err_l
                or "invalid x-api-key" in err_l
                or "error code: 401" in err_l
                or "could not resolve authentication method" in err_l
            ):
                raise AIServiceError("AI authentication failed (invalid API key).")
            raise AIServiceError(f"AI evaluation failed: {str(e)}")

    def _extract_reference_product_name(self, reference_text: str) -> str:
        if not reference_text:
            return ""
        m = re.search(r"(?:product|brand)\s*name\s*[:\-]\s*(.+)", reference_text, flags=re.IGNORECASE)
        if not m:
            return ""
        name = m.group(1).split("\n")[0].strip()
        return re.sub(r"\s+", " ", name)

    def _extract_significant_tokens(self, text: str) -> List[str]:
        stop = {
            "the", "and", "with", "for", "from", "into", "this", "that",
            "premium", "original", "natural", "flavor", "flavour",
            "ice", "cream", "food", "foods"
        }
        tokens = re.findall(r"[a-z0-9]+", (text or "").lower())
        return [t for t in tokens if len(t) >= 3 and t not in stop]

    def _is_strict_reference_row(self, item_l: str) -> bool:
        normalized = re.sub(r"\s+", " ", item_l or "").strip()
        return (
            "according to reference" in normalized
            or "matching according" in normalized
            or ("match" in normalized and "reference" in normalized)
            or ("correct" in normalized and "reference" in normalized)
        )

    def _is_presence_only_row(self, item_l: str) -> bool:
        return ("present" in item_l) and (not self._is_strict_reference_row(item_l))

    def _presence_topic(self, item_l: str) -> str:
        if "product name" in item_l:
            return "product_name"
        if "ingredient" in item_l:
            return "ingredients"
        if "allergen" in item_l:
            return "allergens"
        if "portion size" in item_l:
            return "portion_size"
        if "net weight" in item_l or "weight" in item_l:
            return "net_weight"
        if "shelf life" in item_l or "best before" in item_l:
            return "shelf_life"
        if "claims" in item_l:
            return "claims"
        if "nutrition value" in item_l or "nutrition" in item_l:
            return "nutrition"
        if "company address" in item_l or "address" in item_l:
            return "company_address"
        if "reference intake statement" in item_l:
            return "reference_intake"
        if "ri icon" in item_l:
            return "ri_icon"
        return "generic"

    def _presence_detected(self, topic: str, reference_text: str, artwork_text: str) -> Tuple[bool, str]:
        art_l = (artwork_text or "").lower()

        if topic == "product_name":
            ref_name = self._extract_reference_product_name(reference_text)
            matched_tokens = self._match_reference_tokens_in_artwork(ref_name, artwork_text)
            if (ref_name and ref_name.lower() in art_l) or matched_tokens:
                evidence = ref_name if (ref_name and ref_name.lower() in art_l) else ", ".join(matched_tokens[:4])
                return True, f"Detected product-name evidence: {evidence or 'token match'}."
            return False, "No product-name evidence found."

        if topic == "ingredients":
            if "ingredient" in art_l:
                return True, "Detected ingredients section text."
            return False, "Ingredients section not detected."

        if topic == "allergens":
            if "contains" in art_l and ("milk" in art_l or "soya" in art_l):
                return True, "Detected allergen statement (contains/milk/soya)."
            return False, "Allergen statement not clearly detected."

        if topic == "portion_size":
            if re.search(r"\bserving\b", art_l) and re.search(r"\b\d+\s*x?\s*serving\b|\b\d+\s*ml\b", art_l):
                return True, "Detected serving/portion-size text."
            return False, "Serving/portion-size text not detected."

        if topic == "net_weight":
            if re.search(r"\b\d+\s*ml\b", art_l) and re.search(r"\b\d+\s*g\b", art_l):
                return True, "Detected net-weight values."
            return False, "Net-weight values not clearly detected."

        if topic == "shelf_life":
            if "best before" in art_l:
                return True, "Detected shelf-life statement ('Best before')."
            return False, "Shelf-life statement not detected."

        if topic == "claims":
            if any(k in art_l for k in ("crafted with fresh cream", "plastic", "tub", "dispose")):
                return True, "Detected claims text on artwork."
            return False, "Claims text not clearly detected."

        if topic == "nutrition":
            if "nutrition" in art_l and any(k in art_l for k in ("energy", "fat", "salt")):
                return True, "Detected nutrition table text."
            return False, "Nutrition table text not clearly detected."

        if topic == "company_address":
            if "packed for" in art_l and any(k in art_l for k in ("lane", "park", "town")):
                return True, "Detected company address text."
            return False, "Company address text not clearly detected."

        if topic == "reference_intake":
            if "reference intake" in art_l or ("average adult" in art_l and "8400" in art_l):
                return True, "Detected reference-intake statement text."
            return False, "Reference-intake statement not clearly detected."

        if topic == "ri_icon":
            if "ri icon" in art_l and ("present" in art_l or "shows" in art_l or "visible" in art_l):
                return True, "Detected RI icon presence evidence."
            has_kj = bool(re.search(r"\b\d+\s*kj\b", art_l))
            has_kcal = bool(re.search(r"\b\d+\s*kcal\b", art_l))
            has_pct = bool(re.search(r"\b\d+\s*%\b", art_l))
            has_reference_intake = "reference intake" in art_l
            if has_kj and has_kcal and (has_pct or has_reference_intake):
                return True, "Detected RI icon values (kJ/kcal/%)."
            return False, "RI icon values not clearly detected."

        if art_l.strip():
            return True, "Artwork text detected, but no specific topic matcher was available."
        return False, "Required content not clearly detected."

    def _match_reference_tokens_in_artwork(self, reference_name: str, artwork_text: str) -> List[str]:
        ref_tokens = list(dict.fromkeys(self._extract_significant_tokens(reference_name)))
        art_l = (artwork_text or "").lower()
        return [t for t in ref_tokens if t in art_l]

    def _extract_allergen_terms(self, reference_text: str) -> List[str]:
        """Extract allergen tokens from reference (e.g., MILK, SOYA)."""
        if not reference_text:
            return []
        line_match = re.search(r"contains\s*:\s*([^\n]+)", reference_text, flags=re.IGNORECASE)
        line = line_match.group(1) if line_match else reference_text
        tokens = re.findall(r"\b[A-Z]{2,}\b", line)
        cleaned = [t.strip().strip(".,;:") for t in tokens if t.strip()]
        return list(dict.fromkeys(cleaned))

    def _contains_emphasis_cue(self, text: str) -> bool:
        cues = ("bold", "emphasis", "emphasized", "highlight", "uppercase", "upper case")
        t = (text or "").lower()
        return any(c in t for c in cues)

    def _extract_reference_term_bold_state(self, reference_text: str, term: str) -> Optional[bool]:
        """Return True if term appears bold in reference, False if only plain, None if not found."""
        if not reference_text or not term:
            return None
        term_re = re.escape(term)
        if re.search(rf"\*\*\s*{term_re}\s*\*\*", reference_text, flags=re.IGNORECASE):
            return True
        if re.search(rf"\b{term_re}\b", reference_text, flags=re.IGNORECASE):
            return False
        return None

    def _infer_artwork_term_bold_state(self, notes: str, term: str) -> Optional[bool]:
        """
        Infer bold/unbold statement for a term from AI notes text.
        True=bold, False=plain/unbold, None=unknown.
        """
        if not notes or not term:
            return None
        n = notes.lower()
        t = re.escape(term.lower())
        neg_patterns = [
            rf"{t}[^.:\n]{{0,60}}(not bold|unbold|plain|regular)",
            rf"(not bold|unbold|plain|regular)[^.:\n]{{0,60}}{t}",
        ]
        pos_patterns = [
            rf"{t}[^.:\n]{{0,60}}(bold|emphasized|highlighted)",
            rf"(bold|emphasized|highlighted)[^.:\n]{{0,60}}{t}",
        ]
        for p in neg_patterns:
            if re.search(p, n):
                return False
        for p in pos_patterns:
            if re.search(p, n):
                return True
        # Group-level fallback.
        if "allergen" in n:
            if re.search(r"(allergen[^.:\n]{0,80}(not bold|unbold|plain|regular))|((not bold|unbold|plain|regular)[^.:\n]{0,80}allergen)", n):
                return False
            if re.search(r"(allergen[^.:\n]{0,80}(bold|emphasized|highlighted))|((bold|emphasized|highlighted)[^.:\n]{0,80}allergen)", n):
                return True
        return None

    def _extract_weight_token(self, text: str) -> str:
        if not text:
            return ""
        # Primary pattern: 500 ml /340g e
        m = re.search(r"\b\d+\s*ml\s*/\s*\d+\s*g(?:\s*e)?\b", text, flags=re.IGNORECASE)
        if m:
            return m.group(0).strip()
        # Fallback pattern if OCR drops trailing unit
        m = re.search(r"\b\d+\s*ml\s*/\s*\d+\b", text, flags=re.IGNORECASE)
        return m.group(0).strip() if m else ""

    def _has_space_before_ml(self, token: str) -> bool:
        return bool(re.search(r"\d+\s+ml", token, flags=re.IGNORECASE))

    def _extract_shelf_life_phrase(self, text: str) -> str:
        if not text:
            return ""
        # Keep original case for strict comparison.
        m = re.search(r"(Best\s*before\s*:\s*[^\n\r.;]+)", text, flags=re.IGNORECASE)
        if not m:
            return ""
        return re.sub(r"\s+", " ", m.group(1).strip())

    def _extract_reference_intake_values(self, text: str) -> Dict[str, str]:
        if not text:
            return {}
        segment = text
        marker = re.search(r"reference intake", text, flags=re.IGNORECASE)
        if marker:
            start = max(0, marker.start() - 20)
            end = min(len(text), marker.start() + 260)
            segment = text[start:end]
        kj_match = re.search(r"(\d{3,5})\s*kj", segment, flags=re.IGNORECASE)
        kcal_match = re.search(r"(\d{3,5})\s*kcal", segment, flags=re.IGNORECASE)
        return {
            "segment": re.sub(r"\s+", " ", segment.strip()),
            "kj": kj_match.group(1) if kj_match else "",
            "kcal": kcal_match.group(1) if kcal_match else "",
        }

    def _extract_ri_icon_values(self, text: str) -> Dict[str, str]:
        if not text:
            return {}
        scope = text
        marker = re.search(r"ri icon", text, flags=re.IGNORECASE)
        if marker:
            start = max(0, marker.start() - 20)
            end = min(len(text), marker.start() + 280)
            scope = text[start:end]
        else:
            intake_marker = re.search(r"reference intake", text, flags=re.IGNORECASE)
            if intake_marker:
                start = max(0, intake_marker.start() - 220)
                end = min(len(text), intake_marker.start() + 120)
                scope = text[start:end]
        kj = re.search(r"(\d{2,5})\s*kj", scope, flags=re.IGNORECASE)
        kcal = re.search(r"(\d{2,5})\s*kcal", scope, flags=re.IGNORECASE)
        pcts = re.findall(r"(\d{1,3})\s*%", scope, flags=re.IGNORECASE)
        return {
            "kj": kj.group(1) if kj else "",
            "kcal": kcal.group(1) if kcal else "",
            "pct1": pcts[0] if len(pcts) > 0 else "",
            "pct2": pcts[1] if len(pcts) > 1 else "",
            "scope": re.sub(r"\s+", " ", scope.strip()),
        }

    def _apply_rule_overrides(
        self,
        item: str,
        reference_text: str,
        artwork_text: str,
        result: Dict[str, Any],
    ) -> Dict[str, Any]:
        """
        Deterministic corrections for high-signal product-name rows.
        """
        item_l = re.sub(r"\s+", " ", (item or "").lower()).strip()

        # Global rule requested by product team:
        # - "present" rows should be evaluated only for presence.
        # - strict matching is handled by "according to reference / matching / correct" rows.
        if self._is_presence_only_row(item_l) and "product name" not in item_l:
            topic = self._presence_topic(item_l)
            combined_artwork_view = " ".join(
                [
                    artwork_text or "",
                    str(result.get("artwork_info", "")),
                    str(result.get("comments", "")),
                ]
            )
            is_present, evidence = self._presence_detected(topic, reference_text, combined_artwork_view)
            if is_present:
                result["decision"] = "Compliant"
                result["comments"] = "Yes, it is present on artwork."
                result["artwork_info"] = evidence
            else:
                result["decision"] = "Reject"
                result["comments"] = "No, it is not clearly present on artwork."
                result["artwork_info"] = evidence
            return result

        if "allergen" in item_l and "according to reference" in item_l:
            allergen_terms = self._extract_allergen_terms(reference_text)
            art_upper = (artwork_text or "").upper()
            missing_terms = [t for t in allergen_terms if t not in art_upper]
            combined_ai_notes = " ".join(
                [
                    str(result.get("reference_info", "")),
                    str(result.get("artwork_info", "")),
                    str(result.get("comments", "")),
                ]
            )

            if missing_terms:
                result["decision"] = "Reject"
                result["comments"] = (
                    "Allergen terms do not fully match reference document. "
                    f"Missing in artwork text: {', '.join(missing_terms)}."
                )
                result["reference_info"] = (
                    f"Reference allergen terms: {', '.join(allergen_terms)}."
                    if allergen_terms else result.get("reference_info", "")
                )
                result["artwork_info"] = "Artwork allergen text is missing one or more required terms."
                return result

            # Enforce bidirectional bold/unbold parity from reference <-> artwork.
            mismatched_terms = []
            unverifiable_terms = []
            for term in allergen_terms:
                ref_state = self._extract_reference_term_bold_state(reference_text, term)
                if ref_state is None:
                    continue
                art_state = self._infer_artwork_term_bold_state(combined_ai_notes, term)
                if art_state is None:
                    unverifiable_terms.append(term)
                    continue
                if art_state != ref_state:
                    mismatched_terms.append((term, ref_state, art_state))

            if mismatched_terms:
                mismatch_labels = [
                    f"{term}: reference={'bold' if ref_state else 'plain'}, artwork={'bold' if art_state else 'plain'}"
                    for term, ref_state, art_state in mismatched_terms
                ]
                result["decision"] = "Reject"
                result["comments"] = (
                    "Allergen formatting mismatch detected (bold/unbold parity failed): "
                    + "; ".join(mismatch_labels)
                )
                result["reference_info"] = (
                    f"Reference allergen terms: {', '.join(allergen_terms)}."
                    if allergen_terms else result.get("reference_info", "")
                )
                result["artwork_info"] = "Artwork allergen formatting does not match reference formatting."
                return result

            if unverifiable_terms:
                result["decision"] = "Reject"
                result["comments"] = (
                    "Allergen terms are present but bold/unbold formatting could not be verified for: "
                    + ", ".join(unverifiable_terms)
                )
                if allergen_terms:
                    result["reference_info"] = f"Reference allergen terms requiring emphasis: {', '.join(allergen_terms)}."
                result["artwork_info"] = "Artwork allergen terms detected, but formatting parity is not verifiable."
                return result

        if "net weight" in item_l and "according to reference" in item_l:
            reference_weight = self._extract_weight_token(reference_text)
            actual_weight = self._extract_weight_token(
                " ".join(
                    [
                        artwork_text or "",
                        str(result.get("artwork_info", "")),
                        str(result.get("comments", "")),
                    ]
                )
            )

            if reference_weight and actual_weight:
                expected_ml_spacing = self._has_space_before_ml(reference_weight)
                actual_ml_spacing = self._has_space_before_ml(actual_weight)
                normalized_reference = re.sub(r"\s+", " ", reference_weight.strip().lower())
                normalized_actual = re.sub(r"\s+", " ", actual_weight.strip().lower())

                if expected_ml_spacing != actual_ml_spacing or normalized_reference != normalized_actual:
                    result["decision"] = "Reject"
                    result["reference_info"] = f"Reference net weight format: {reference_weight}"
                    result["artwork_info"] = f"Artwork net weight format: {actual_weight}"
                    result["comments"] = (
                        "Net weight formatting does not match reference exactly (strict spacing/format mode)."
                    )
                    return result

                result["decision"] = "Compliant"
                result["reference_info"] = f"Reference net weight format: {reference_weight}"
                result["artwork_info"] = f"Artwork net weight format: {actual_weight}"
                result["comments"] = "Net weight matches reference exactly, including formatting."
                return result

            if reference_weight and not actual_weight:
                result["decision"] = "Reject"
                result["reference_info"] = f"Reference net weight format: {reference_weight}"
                result["artwork_info"] = "Unable to find a full net weight token in artwork text."
                result["comments"] = "Net weight cannot be verified against reference in strict format mode."
                return result

        if "shelf life" in item_l and "according to reference" in item_l:
            reference_phrase = self._extract_shelf_life_phrase(reference_text)
            artwork_phrase = self._extract_shelf_life_phrase(
                " ".join(
                    [
                        artwork_text or "",
                        str(result.get("artwork_info", "")),
                        str(result.get("comments", "")),
                    ]
                )
            )

            if reference_phrase and artwork_phrase:
                if reference_phrase != artwork_phrase:
                    result["decision"] = "Reject"
                    result["reference_info"] = f"Reference shelf-life text: {reference_phrase}"
                    result["artwork_info"] = f"Artwork shelf-life text: {artwork_phrase}"
                    result["comments"] = "Shelf-life text case/format does not exactly match reference."
                    return result

                result["decision"] = "Compliant"
                result["reference_info"] = f"Reference shelf-life text: {reference_phrase}"
                result["artwork_info"] = f"Artwork shelf-life text: {artwork_phrase}"
                result["comments"] = "Shelf-life text matches reference exactly, including case."
                return result

            if reference_phrase and not artwork_phrase:
                result["decision"] = "Reject"
                result["reference_info"] = f"Reference shelf-life text: {reference_phrase}"
                result["artwork_info"] = "Unable to extract shelf-life text from artwork."
                result["comments"] = "Shelf-life cannot be verified in strict mode."
                return result

        if "reference intake statement" in item_l and "according to reference" in item_l:
            ref_vals = self._extract_reference_intake_values(reference_text)
            art_vals = self._extract_reference_intake_values(
                " ".join(
                    [
                        artwork_text or "",
                        str(result.get("artwork_info", "")),
                        str(result.get("comments", "")),
                    ]
                )
            )
            if ref_vals.get("kj") and ref_vals.get("kcal"):
                if (ref_vals.get("kj") == art_vals.get("kj")) and (ref_vals.get("kcal") == art_vals.get("kcal")):
                    result["decision"] = "Compliant"
                    result["reference_info"] = (
                        f"Reference intake values: {ref_vals.get('kj')} kJ / {ref_vals.get('kcal')} kcal"
                    )
                    result["artwork_info"] = (
                        f"Artwork intake values: {art_vals.get('kj') or '?'} kJ / {art_vals.get('kcal') or '?'} kcal"
                    )
                    result["comments"] = "Reference intake statement values match reference exactly."
                    return result

                result["decision"] = "Reject"
                result["reference_info"] = (
                    f"Reference intake values: {ref_vals.get('kj')} kJ / {ref_vals.get('kcal')} kcal"
                )
                result["artwork_info"] = (
                    f"Artwork intake values: {art_vals.get('kj') or '?'} kJ / {art_vals.get('kcal') or '?'} kcal"
                )
                result["comments"] = "Reference intake statement values do not match reference exactly."
                return result

        if "ri icon" in item_l and self._is_strict_reference_row(item_l):
            ref_vals = self._extract_ri_icon_values(reference_text)
            art_vals = self._extract_ri_icon_values(
                " ".join(
                    [
                        artwork_text or "",
                        str(result.get("artwork_info", "")),
                        str(result.get("comments", "")),
                    ]
                )
            )
            if ref_vals.get("kj") and ref_vals.get("kcal") and ref_vals.get("pct1") and ref_vals.get("pct2"):
                ref_tuple = (ref_vals["kj"], ref_vals["kcal"], ref_vals["pct1"], ref_vals["pct2"])
                art_tuple = (art_vals.get("kj", ""), art_vals.get("kcal", ""), art_vals.get("pct1", ""), art_vals.get("pct2", ""))
                if ref_tuple == art_tuple:
                    result["decision"] = "Compliant"
                    result["reference_info"] = (
                        f"Reference RI icon values: {ref_vals['kj']} kJ, {ref_vals['kcal']} kcal, {ref_vals['pct1']}%, {ref_vals['pct2']}%"
                    )
                    result["artwork_info"] = (
                        f"Artwork RI icon values: {art_tuple[0]} kJ, {art_tuple[1]} kcal, {art_tuple[2]}%, {art_tuple[3]}%"
                    )
                    result["comments"] = "RI icon values match reference exactly."
                    return result
                result["decision"] = "Reject"
                result["reference_info"] = (
                    f"Reference RI icon values: {ref_vals['kj']} kJ, {ref_vals['kcal']} kcal, {ref_vals['pct1']}%, {ref_vals['pct2']}%"
                )
                result["artwork_info"] = (
                    f"Artwork RI icon values: {art_tuple[0] or '?'} kJ, {art_tuple[1] or '?'} kcal, {art_tuple[2] or '?'}%, {art_tuple[3] or '?'}%"
                )
                result["comments"] = "RI icon values do not match reference exactly."
                return result

        if "nutrition" in item_l and "present" in item_l and "according to reference" not in item_l:
            art_l = (artwork_text or "").lower()
            has_nutrition_section = (
                ("nutrition" in art_l and ("energy" in art_l or "fat" in art_l or "salt" in art_l))
                or ("nutritional values" in art_l)
            )
            if has_nutrition_section:
                result["decision"] = "Compliant"
                result["comments"] = "Nutrition information presence confirmed on artwork."
                if not result.get("artwork_info"):
                    result["artwork_info"] = "Nutrition panel text detected (e.g., energy/fat/salt)."
            else:
                result["decision"] = "Reject"
                result["comments"] = "Nutrition information panel is not clearly detected on artwork."
            return result

        if "product name" not in item_l:
            return result

        art_l = (artwork_text or "").lower()
        ref_name = self._extract_reference_product_name(reference_text)
        ref_name_l = ref_name.lower()
        matched_tokens = self._match_reference_tokens_in_artwork(ref_name, artwork_text)
        token_hits = len(matched_tokens)
        exact_match = bool(ref_name_l and ref_name_l in art_l)

        if "present" in item_l and "according to reference" not in item_l:
            if exact_match or token_hits >= 1:
                result["decision"] = "Compliant"
                result["comments"] = "Product name presence confirmed from artwork text."
                if ref_name:
                    result["reference_info"] = f"Reference product name: {ref_name}"
                if exact_match:
                    result["artwork_info"] = "Detected full reference product name string in artwork text."
                else:
                    result["artwork_info"] = f"Detected reference product-name tokens in artwork: {', '.join(matched_tokens)}."
            return result

        if "according to reference" in item_l or "match" in item_l:
            if exact_match or token_hits >= 2:
                result["decision"] = "Compliant"
                result["comments"] = "Product name matches reference document."
                if ref_name:
                    result["reference_info"] = f"Reference product name: {ref_name}"
                if exact_match:
                    result["artwork_info"] = "Artwork product name exactly matches reference product name."
                else:
                    result["artwork_info"] = f"Artwork product-name tokens align with reference: {', '.join(matched_tokens)}."
            return result

        return result

    def _parse_model_response(self, response_text: str, item: str, allow_salvage: bool = True) -> Dict[str, Any]:
        """
        Parse model JSON robustly.
        Handles fenced output, stray text, unescaped newlines, and partial malformed JSON.
        """
        if not response_text or not response_text.strip():
            print(f"Warning: Empty response for item: {item}")
            return self._normalize_result({"decision": "N/A", "comments": "Empty response"})

        candidate = self._extract_json_object(response_text)

        # Try parsing extracted JSON, then repaired JSON.
        attempts = [candidate, self._repair_json_text(candidate)]
        seen = set()
        for attempt in attempts:
            if not attempt:
                continue
            if attempt in seen:
                continue
            seen.add(attempt)
            try:
                parsed = json.loads(attempt)
                return self._normalize_result(parsed)
            except json.JSONDecodeError as e:
                print(f"Debug: JSON parse attempt failed - {str(e)[:100]}")
                continue

        if not allow_salvage:
            raise json.JSONDecodeError(
                "Unable to parse model response without salvage",
                candidate or response_text or "",
                0,
            )

        # If strict parsing still fails, do conservative salvage and mark as N/A.
        salvaged = self._salvage_result_fields(candidate or response_text)
        if salvaged:
            salvaged["decision"] = "N/A"
            salvaged["comments"] = (
                "AI response was malformed; manual review recommended. "
                + (salvaged.get("comments") or "").strip()
            ).strip()
            print(f"Info: Successfully salvaged fields from malformed response")
            return self._normalize_result(salvaged)

        print(f"Error: Unable to parse model response for item: {item}")
        print(f"Debug: Raw response (first 200 chars): {response_text[:200]}")
        raise json.JSONDecodeError(
            "Unable to parse model response",
            candidate or response_text or "",
            0,
        )

    def _extract_json_object(self, text: str) -> str:
        """Extract the first balanced JSON object from model output."""
        if not text:
            return "{}"

        cleaned = text.strip()
        if cleaned.startswith("```"):
            cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
            cleaned = re.sub(r"\s*```$", "", cleaned)

        start = cleaned.find("{")
        if start == -1:
            return cleaned

        depth = 0
        in_string = False
        escaped = False
        end = -1
        for i in range(start, len(cleaned)):
            ch = cleaned[i]
            if escaped:
                escaped = False
                continue
            if ch == "\\":
                escaped = True
                continue
            if ch == '"':
                in_string = not in_string
                continue
            if in_string:
                continue
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    end = i
                    break

        if end != -1:
            return cleaned[start:end + 1]

        # Fallback for truncated objects
        return cleaned[start:]

    def _repair_json_text(self, text: str) -> str:
        """Apply small safe repairs for common malformed JSON patterns."""
        if not text:
            return text

        repaired = text

        # Replace smart quotes with regular quotes
        repaired = repaired.replace("\u201c", '"').replace("\u201d", '"').replace("\u2018", "'").replace("\u2019", "'")

        # Normalize line endings
        repaired = repaired.replace("\r\n", "\n").replace("\r", "\n")

        # Remove control characters but keep newlines
        repaired = "".join(ch for ch in repaired if ch == "\n" or ord(ch) >= 32)

        # Escape literal newlines/tabs within strings
        out = []
        in_string = False
        escaped = False
        for i, ch in enumerate(repaired):
            if escaped:
                out.append(ch)
                escaped = False
                continue
            if ch == "\\":
                out.append(ch)
                escaped = True
                continue
            if ch == '"':
                out.append(ch)
                in_string = not in_string
                continue
            if in_string and ch == "\n":
                out.append("\\n")
                continue
            if in_string and ch == "\t":
                out.append("\\t")
                continue
            out.append(ch)
        repaired = "".join(out)

        # Escape unescaped inner quotes inside string values.
        out = []
        in_string = False
        escaped = False
        for i, ch in enumerate(repaired):
            if escaped:
                out.append(ch)
                escaped = False
                continue
            if ch == "\\":
                out.append(ch)
                escaped = True
                continue
            if ch == '"':
                if not in_string:
                    in_string = True
                    out.append(ch)
                    continue

                # Decide if this is a closing quote for key/value or an inner quote.
                j = i + 1
                while j < len(repaired) and repaired[j].isspace():
                    j += 1
                if j < len(repaired) and repaired[j] in {",", "}", "]", ":"}:
                    in_string = False
                    out.append(ch)
                else:
                    out.append('\\"')
                continue
            out.append(ch)
        repaired = "".join(out)
        if in_string:
            repaired += '"'

        # Balance braces/brackets for truncated JSON output.
        brace_depth = 0
        bracket_depth = 0
        in_string = False
        escaped = False
        for ch in repaired:
            if escaped:
                escaped = False
                continue
            if ch == "\\":
                escaped = True
                continue
            if ch == '"':
                in_string = not in_string
                continue
            if in_string:
                continue
            if ch == "{":
                brace_depth += 1
            elif ch == "}":
                brace_depth = max(0, brace_depth - 1)
            elif ch == "[":
                bracket_depth += 1
            elif ch == "]":
                bracket_depth = max(0, bracket_depth - 1)
        if bracket_depth:
            repaired += "]" * bracket_depth
        if brace_depth:
            repaired += "}" * brace_depth

        # Remove trailing commas before object/array close
        repaired = re.sub(r",\s*([}\]])", r"\1", repaired)

        # Fix common JSON issues
        repaired = re.sub(r':\s*"([^"]*?)\s*"([,}\]])', r': "\1"\2', repaired)  # Extra spaces in strings

        return repaired

    def _salvage_result_fields(self, text: str) -> Optional[Dict[str, Any]]:
        """
        Best-effort extraction when JSON is malformed.
        """
        if not text:
            return None

        decision_match = re.search(
            r'"?decision"?\s*:\s*"?(Compliant|Reject|N/?A|Complaint|Fail(?:ed)?|Pass(?:ed)?|Non-?Compliant)"?',
            text,
            flags=re.IGNORECASE,
        )
        decision = decision_match.group(1) if decision_match else "N/A"
        if decision.lower() == "complaint":
            decision = "Compliant"

        def _extract_field(field: str) -> str:
            pattern = (
                rf'"?{field}"?\s*:\s*'
                rf'("(?P<quoted>(?:\\.|[^"\\])*)"|(?P<bare>.*?))'
                rf'(?=,\s*"?(?:decision|reference_info|artwork_info|comments)"?\s*:|[}}]|\Z)'
            )
            m = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
            if not m:
                return ""
            val = m.group("quoted") if m.group("quoted") is not None else (m.group("bare") or "")
            return val.strip().strip(",").strip()

        result = {
            "decision": decision,
            "reference_info": _extract_field("reference_info"),
            "artwork_info": _extract_field("artwork_info"),
            "comments": _extract_field("comments"),
        }

        if not any(result.values()):
            return None
        return result

    def _normalize_result(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """Normalize model/fallback output to expected schema."""
        decision_raw = str(result.get("decision", "")).strip().lower()
        mapping = {
            "compliant": "Compliant",
            "reject": "Reject",
            "n/a": "N/A",
            "na": "N/A",
            "complaint": "Compliant",
            "pass": "Compliant",
            "passed": "Compliant",
            "approve": "Compliant",
            "approved": "Compliant",
            "fail": "Reject",
            "failed": "Reject",
            "non-compliant": "Reject",
            "noncompliant": "Reject",
        }
        decision = mapping.get(decision_raw, "N/A")
        return {
            "decision": decision,
            "reference_info": str(result.get("reference_info", "")).strip(),
            "artwork_info": str(result.get("artwork_info", "")).strip(),
            "comments": str(result.get("comments", "")).strip(),
        }

    def _fallback_evaluation(self, item: str, reference_text: str, artwork_text: str) -> Dict[str, Any]:
        """
        Lightweight heuristic evaluation when Anthropic API key is unavailable.
        Keeps the pipeline functional for local testing and Airtable updates.
        """
        item_l = (item or "").lower()
        art_l = (artwork_text or "").lower()
        ref_snippet = (reference_text or "").strip().replace("\n", " ")[:220] or "Reference requirement unavailable"
        art_snippet = (artwork_text or "").strip().replace("\n", " ")[:220] or "No OCR text detected"

        if not art_l.strip():
            decision = "N/A"
            comments = "No OCR text detected from artwork; manual review required."
        elif "product name" in item_l:
            decision = "Compliant" if ("product" in art_l and "label" in art_l) else "Reject"
            comments = "Fallback check: looked for product label/name tokens."
        elif "net weight" in item_l or "weight" in item_l:
            has_weight = bool(re.search(r"\b\d+(?:\.\d+)?\s?(?:g|kg|ml|l)\b", art_l))
            decision = "Compliant" if has_weight else "Reject"
            comments = "Fallback check: looked for numeric metric unit patterns."
        elif "ingredient" in item_l:
            decision = "Compliant" if "ingredient" in art_l else "Reject"
            comments = "Fallback check: looked for ingredients text."
        elif "expiry" in item_l or "expir" in item_l or "date" in item_l:
            has_date = bool(re.search(r"\b(?:0[1-9]|1[0-2])[\-/](?:19|20)?\d{2}\b", art_l)) or bool(re.search(r"\b\d{2}[\-/]\d{2}[\-/]\d{2,4}\b", art_l))
            decision = "Compliant" if has_date else "Reject"
            comments = "Fallback check: looked for date-like pattern."
        elif "logo" in item_l:
            decision = "N/A"
            comments = "Fallback evaluator cannot reliably verify logo presence from OCR text only."
        else:
            tokens = [t for t in re.findall(r"[a-z]{4,}", item_l) if t not in {"must", "be", "with", "from", "that", "this", "have", "appear", "visible", "clearly", "displayed"}]
            overlap = sum(1 for t in tokens if t in art_l)
            decision = "Compliant" if overlap >= max(1, len(tokens) // 2) else "Reject"
            comments = "Fallback check: keyword overlap against OCR text."

        fallback_result = {
            "checklist_item": item,
            "decision": decision,
            "reference_info": ref_snippet,
            "artwork_info": art_snippet,
            "comments": comments,
        }
        return self._apply_rule_overrides(item, reference_text, artwork_text, fallback_result)

    def run_compliance_check(self, checklist_path: str, reference_path: str,
                            artwork_path: str) -> Dict[str, Any]:
        """
        Run full compliance check and return results with summary.
        """
        if not self.force_fallback and (not self.ai_enabled or not self.client):
            raise AIServiceError("Primary AI evaluation is unavailable (missing ANTHROPIC_API_KEY).")

        # Parse inputs
        checklist_items = self.parse_checklist(checklist_path)
        reference_text = self.extract_reference_text(reference_path)
        artwork_text = self.extract_artwork_text(artwork_path)

        # Evaluate each item
        results = []
        for item in checklist_items:
            result = self.evaluate_checklist_item(item, reference_text, artwork_text, artwork_path=artwork_path)
            results.append(result)

        # Build summary
        summary = self._build_summary(results)

        return {
            'summary': summary,
            'results': results
        }

    def _build_summary(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Build compliance summary statistics."""
        total = len(results)
        compliant = sum(1 for r in results if r.get('decision') == 'Compliant')
        reject = sum(1 for r in results if r.get('decision') == 'Reject')
        na = sum(1 for r in results if r.get('decision') == 'N/A')

        compliance_percentage = (compliant / total * 100) if total > 0 else 0

        return {
            'total': total,
            'compliant': compliant,
            'reject': reject,
            'na': na,
            'compliance_percentage': round(compliance_percentage, 1)
        }
